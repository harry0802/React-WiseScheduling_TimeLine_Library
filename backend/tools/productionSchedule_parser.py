from math import ceil
import os
import sys, getopt
import logging
import requests
from openpyxl import load_workbook
from datetime import datetime, date

# parse data from excel file, and insert into database
debug = False
table_name = 'productionSchedule'
ServerName = 'http://localhost:5000'
api_url = f"{ServerName}/api/productionSchedule"
column_names = [
    # 'id',
    'productionArea',
    'machineSN',
    'serialNumber',
    'workOrderSN',
    'productSN',
    'productName',
    'workOrderQuantity',
    'workOrderDate',
    'moldingSecond',
    'hourlyCapacity',
    'dailyCapacity',
    'planOnMachineDate',
    'actualOnMachineDate',
    'moldWorkDays',
    'workDays',
    'planFinishDate',
    'actualFinishDate',
    'comment',
    'dailyWorkingHours',
    'moldCavity',
    'week',
    'singleOrDoubleColor',
    'conversionRate',
    'status',
]

type_mapping = {
    "productionArea": "string",
    "machineSN": "string",
    "serialNumber": "string",
    "workOrderSN": "string",
    "productSN": "string",
    "productName": "string",
    "workOrderQuantity": "int",
    "workOrderDate": "date",
    "moldingSecond": "int",
    "hourlyCapacity": "float",
    "dailyCapacity": "float",
    "planOnMachineDate": "date",
    "actualOnMachineDate": "date",
    "moldWorkDays": "int",
    "workDays": "int",
    "planFinishDate": "date",
    "actualFinishDate": "date",
    "comment": "string",
    "dailyWorkingHours": "int",
    "moldCavity": "int",
    "week": "int",
    "singleOrDoubleColor": "string",
    "conversionRate": "float",
    "status": "string",
}
#-----------------------------------------------------------------------------------
def type_transform(obj):
    format = '%Y/%m/%d'
    for key, value in obj.items():
        if key in type_mapping:
            if type_mapping[key] == "int":
                obj[key] = int(value) if value is not None else None
            elif type_mapping[key] == "float":
                obj[key] = float(value) if value is not None else None
            elif type_mapping[key] == "date":
                obj[key] = datetime.strptime(value, format) if (
                    type(value) is str
                    and value is not None
                    ) else value

def main():
    global debug
    success = 0
    fail = 0
    #### main procedure start here ####
    # get cli parameters
    file_name = sys.argv[1]
    sheet_idx = sys.argv[2]
    wb = load_workbook(file_name, data_only=True)
    sheet = wb.worksheets[int(sheet_idx)]
    #sheet = wb[sheet_name]
    # get data from excel
    for row_ori in sheet.iter_rows(min_row=5, max_col=24, values_only=True):
        # column_name and row mapping
        row = dict(zip(column_names, row_ori))
        if debug:
            print(f"~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
            print(f"row_ori: {row_ori}")
        # if row['productionArea'] is None:
        #     continue
        #default values
        # row['productionArea']= '+++++++++' if row['productionArea'] is None else row['productionArea']
        # row['moldWorkDays'] = -999999 if row['moldWorkDays'] is None else row['moldWorkDays']
        # row['status'] = '@@@@@@@@@@@@' if row['status'] is None else row['status']
        # convert data type
        type_transform(row)
        # convert data format
        row['workOrderDate'] = row['workOrderDate'].strftime('%Y-%m-%d') if row['workOrderDate'] is not None else None
        row['planOnMachineDate'] = row['planOnMachineDate'].strftime('%Y-%m-%d') if row['planOnMachineDate'] is not None else None
        row['planFinishDate'] = row['planFinishDate'].strftime('%Y-%m-%d') if row['planFinishDate'] is not None else None
        row['actualFinishDate'] = row['actualFinishDate'].strftime('%Y-%m-%d') if row['actualFinishDate'] is not None else None
    
        # convert value
        row['workDays'] = ceil(row['workDays']) if row['workDays'] is not None else None
        row['moldWorkDays'] = ceil(row['moldWorkDays']) if row['moldWorkDays'] is not None else None
        row['moldingSecond'] = ceil(row['moldingSecond']) if row['moldingSecond'] is not None else None
        row['hourlyCapacity'] = round(row['hourlyCapacity']) if row['hourlyCapacity'] is not None else None
        row['dailyCapacity'] = round(row['dailyCapacity']) if row['dailyCapacity'] is not None else None
        row.pop('serialNumber') # remove serialNumber, system generated
        # remove empty columns
        empty_cols = [key for key, value in row.items() if value is None ]
        for key in empty_cols:
            row.pop(key, None)
        ret = requests.post(api_url, json=row)
        if debug:
            print(f"~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
            print(ret.json())
            print("@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")
        if ret.json().get('status'):
            success += 1
        else:
            fail += 1
    print(f"success: {success}, fail: {fail}")

def usage():
    print('usage: {} [-h] file_name sheet_idx'.format(sys.argv[0]))
    print('    -h, --help       help information')
    # print('    -d, --debug      force to print debug message')
    # print('    -v, --verbose    force print verbose information')


if __name__ == '__main__':
    try:
        opts, args = getopt.getopt(sys.argv[1:], 'h', ['help'])
    except(getopt.GetoptError):
        print('usage: {} [-h] '.format(sys.argv[0]))
        sys.exit()
    for opt, arg in opts:
        if opt in ("-h", "--help"):
            usage()
            sys.exit()
    if len(args) < 2:
        usage()
        sys.exit()
    main()
